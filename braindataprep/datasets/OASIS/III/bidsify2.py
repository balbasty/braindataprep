"""
Expected input
--------------
OASIS-3/
    sourcedata/
        OASIS3_data_files/
            # demographics + cognitive tests
            demo.tar.gz
            dictionaries.tar.gz
            # Modality-specific BIDS json
            MRI-json.tar.gz
            CT-json.tar.gz
            PET-json.tar.gz
            # derivatives
            FS.tar.gz
            PUP.tar.gz
            # UDS
            UDSa1.tar.gz
            UDSa2.tar.gz
            UDSa3.tar.gz
            UDSa4d.tar.gz
            UDSa4g.tar.gz
            UDSa5.tar.gz
            UDSb1.tar.gz
            UDSb2.tar.gz
            UDSb3.tar.gz
            UDSb4.tar.gz
            UDSb5.tar.gz
            UDSb6.tar.gz
            UDSb7.tar.gz
            UDSb8.tar.gz
            UDSb9.tar.gz
            pychometrics.tar.gz  (== UDSc1)
            UDSd1.tar.gz
            UDSd2.tar.gz
        OAS3{04d}_MR_d{04d}/
            anat{d}.tar.gz
            func{d}.tar.gz
            fmap{d}.tar.gz
            dwi{d}.tar.gz
            swi{d}.tar.gz
        OAS3{04d}_CT_d{04d}/
            CT{d}.tar.gz
        OAS3{04d}_PIB_d{04d}/
            pet{d}.tar.gz
        OAS3{04d}_AV45_d{04d}/
            pet{d}.tar.gz

Expected output
---------------
OASIS-3/
    dataset_description.json
    participants.{tsv|json}
    sessions.json
    phenotypes/
        UDSv2_a1_demographics.{tsv|json}
        UDSv2_a2_informant.{tsv|json}
        UDSv2_a3_family_history.{tsv|json}
        UDSv2_a4_medications.{tsv|json}
        UDSv2_a5_health_history.{tsv|json}
        UDSv2_b1_physical.{tsv|json}
        UDSv2_b2_hiscvd.{tsv|json}
        UDSv2_b3_updrs.{tsv|json}
        UDSv2_b4_cdr.{tsv|json}
        UDSv2_b5_npiq.{tsv|json}
        UDSv2_b6_gds.{tsv|json}
        UDSv2_b7_fas.{tsv|json}
        UDSv2_b8_neurofind.{tsv|json}
        UDSv2_b9_symptoms.{tsv|json}
        UDSv2_c1_neuropsy.{tsv|json}
        UDSv2_d1_diagnosis.{tsv|json}
        UDSv2_d2_medical_conditions.{tsv|json}
    rawdata/
        sub-{04d}/
            sub-{04d}_sessions.tsv
            ses-{d}/
                anat/
                    sub-{:04d}_ses-{:04d}_T1w.{nii.gz|json}
                    sub-{:04d}_ses-{:04d}_T2w.{nii.gz|json}
                    sub-{:04d}_ses-{:04d}_acq-TSE_T2w.{nii.gz|json}
                    sub-{:04d}_ses-{:04d}_FLAIR.{nii.gz|json}
                    sub-{:04d}_ses-{:04d}_T2star.{nii.gz|json}
                    sub-{:04d}_ses-{:04d}_angio.{nii.gz|json}
                perf/
                    sub-{:04d}_ses-{:04d}_pasl.nii.gz
                func/
                    sub-{:04d}_ses-{:04d}_task-rest*_run-{:02d}_bold.nii.gz
                fmap/
                    sub-{:04d}_ses-{:04d}_echo-1_run-01_fieldmap.nii.gz
                dwi/
                    sub-{:04d}_ses-{:04d}_run-{:02d}_dwi.nii.gz
                swi/
                    ...
                pet/
                    sub-{:04d}_trc-PIB_pet.{nii.gz|json}
                    sub-{:04d}_trc-AV45_pet.{nii.gz|json}
"""
import os
import sys
from braindataprep.ioutils import copy_from_buffer, copy_json, fileparts, read_json, write_from_buffer, write_json, write_tsv
import cyclopts
import tarfile
import logging
from pathlib import Path, PosixPath
from typing import Optional, Iterable
from braindataprep.utils.path import (
    get_tree_path,
)
from .keys import allkeys, compat_keys

app = cyclopts.App()
logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    logger.error(
        'Cannot find `openpyxl`. Did you install with [oasis2] flag? '
        'Try `pip install braindataprep[oasis2]`.'
    )

"""Folder containing template README/JSON/..."""
TPLDIR = Path(__file__).parent / 'templates'


@app.default
def bidsify(
    path: Optional[str] = None,
    *,
    keys: Optional[Iterable[str]] = None,
    subs: Iterable[int] = None,
    log: Optional[str] = None,
):
    """
    Convert sourcedata into a BIDS-compliant rawdata directory

    Available keys:
    * raw :                 All the raw imaging data
        * mri :             All the MRI data
            * anat :        All the anatomical MRI data
                * T1w
                * T2w
                * TSE
                * FLAIR
                * T2star
                * angio
            * func :        All fthe functional MRI data
                * pasl
                * asl
                * bold
            * fmap :        All field maps
            * dwi :         All diffusion-weighted MRI data
            * swi :         All susceptibility-weighted MRI data
        * pet :             All the PET data
            * fdg
            * pib
            * av45
            * av1451
        * ct                All the CT data
    * derivatives :         All derivatives
        * fs :              Freesurfer derivatives
        * pup :             PET derivatives
    * meta :                All metadata
        * pheno :           Phenotypes

    Parameters
    ----------
    path : str
        Path to root of all datasets. An `IXI/sourcedata` folder must exist.
    keys : [list of] str
        Only bidsify these keys [default: all]
    subs : [list of] int
        Only bidsify these subjects [default: all]
    log : str
        Path to log file
    """
    if isinstance(log, str):
        logging.basicConfig(filename=log)
        log = logger
    else:
        log = log or logger
    logger.info('OASIS-III - bidsify')
    path = Path(get_tree_path(path))
    keys = set(keys or allkeys)
    root = path / 'OASIS-3'
    src = root / 'sourcedata'
    raw = root / 'rawdata'
    pheno = root / 'phenotypes'
    if not src.exists():
        raise FileNotFoundError('sourcedata folder not found')

    # ------------------------------------------------------------------
    #   Write toplevel meta-information
    #    - README
    #    - dataset_description.json
    #    - participants.tsv
    #    - participants.json
    # ------------------------------------------------------------------
    if keys.intersection(compat_keys("meta")):

        copy_from_buffer(
            TPLDIR / 'README',
            root / 'README',
            logger=logger,
        )

        copy_json(
            TPLDIR / 'dataset_description.json',
            root / 'dataset_description.json',
            logger=logger,
        )

        copy_json(
            TPLDIR / 'participants.json',
            root / 'participants.json',
            logger=logger,
        )

        json_pheno = TPLDIR / 'phenotypes' / 'UDSv2'
        for fname in json_pheno.glob('*.json'):
            ofname = pheno / fname.name
            if 'a3' in fname.name:
                obj = read_json(fname)
                for key, val in dict(obj).items():
                    if "{d}" in key:
                        assert key.startswith(("SIB", "KID", "REL")), key
                        nb = 20 if key.startswith("SIB") else 15
                        del obj[key]
                        desc = val.pop("Description")
                        for d in range(nb):
                            obj[key.format(d=d)] = {
                                'Description': desc.format(d=d),
                                **val
                            }
                write_json(obj, ofname, logger=logger)
            else:
                copy_json(fname, ofname, logger=logger)

        # make_phenotypes()

        # make_participants(
        #     pheno / 'UDSv2_a1_demographics.tsv',
        #     root / 'participants.tsv',
        # )

        copy_json(
            TPLDIR / 'sessions.json',
            root / 'sessions.json',
            logger=logger,
        )

    # we always need this to write subject-specific session files
    # session_tables = make_sessions(
    #     [pheno / 'UDSv2_a1_demographics.tsv',
    #      pheno / 'UDSv2_b4_cdr.tsv',
    #      pheno / 'UDSv2_d1_diagnosis.tsv'],
    #     src / 'oasis_longitudinal_demographics.xlsx'
    # )

    # ------------------------------------------------------------------
    #   Convert raw and minimally processed data
    # ------------------------------------------------------------------

    all_sessions = list(sorted(src.glob('OAS3*')))
    for session in all_sessions:

        participant_id, modality_type, session_id \
            = session.name.split('_')
        participant_id = int(participant_id[4:])
        session_id = int(session_id[1:])

        if subs and participant_id not in subs:
            logger.info(f'skip sub-{participant_id:04d}')
            continue

        if modality_type == 'MR':
            if not keys.intersection(compat_keys("mri")):
                continue
        elif modality_type == "CT":
            if not keys.intersection(compat_keys("ct")):
                continue
        elif modality_type == "FDG":
            if not keys.intersection(compat_keys("fdg")):
                continue
        elif modality_type == "PIB":
            if not keys.intersection(compat_keys("pib")):
                continue
        elif modality_type == "AV45":
            if not keys.intersection(compat_keys("av45")):
                continue
        else:
            continue

        session_path = (
            raw / f'sub-{participant_id:04d}' / f'ses-{session_id:04d}'
        )
        all_scans = list(sorted(session.glob('*.tar.gz')))
        for tarpath in all_scans:
            submodality_type = fileparts(tarpath)[1][:-1]
            modality_path = os.path.join(session_path, submodality_type)

            if not keys.intersection(compat_keys(submodality_type)):
                continue

            with tarfile.open(tarpath, 'r:gz') as f:
                niimember = None
                jsonmember = None
                bvalmember = None
                bvecmember = None
                tsvmember = None
                for member in f.getmembers():
                    if niimember and jsonmember:
                        break
                    if member.name.endswith('.nii.gz'):
                        niimember = member
                    if member.name.endswith('.json'):
                        jsonmember = member
                    if member.name.endswith('.bval'):
                        bvalmember = member
                    if member.name.endswith('.bvec'):
                        bvecmember = member
                    if member.name.endswith('.tsv'):
                        tsvmember = member

                if niimember is None:
                    logger.warning(
                        f"No nifti found in {tarpath.name}"
                    )
                    continue

                find_key = None
                for key in ("_T1w", "_T2w", "TSE", "_FLAIR", "_T2star",
                            "angio", "asl", "bold", "fmap", "_dwi", "_swi",
                            "_ct", "fdg", "pib", "av45", "av1451"):
                    if key in niimember.name.lower():
                        find_key = key.split('_')[-1]
                if find_key and not keys.intersection(compat_keys(find_key)):
                    logger.info(f'skip {PosixPath(niimember.name).name}')

                basename = niimember.name.split('/')[-1].split('.')[0]
                flags = list(basename.split('_'))

                # fix subdirectory for perfusion
                if basename.endswith('asl'):
                    submodality_type = 'perf'
                    modality_path = session_path / 'perf'

                # fix naming convention for fieldmaps
                if basename.endswith('fieldmap'):
                    necho = None
                    iecho = None
                    irun = None
                    for i, flag in enumerate(flags):
                        if flag.startswith('echo-'):
                            iecho = i
                            necho = int(flag.split('-')[-1])
                        if flag.startswith('run-'):
                            irun = i
                    if iecho is not None:
                        flags[-1] = f'magnitude{necho:d}'
                        if irun > iecho:
                            del flags[irun]
                            del flags[iecho]
                        else:
                            del flags[iecho]
                            del flags[irun]
                    else:
                        flags[-1] = 'phasediff'
                        del flags[irun]

                # fix naming convention for PET tracers
                if basename.endswith('pet'):
                    for i, flag in enumerate(flags):
                        if flag.startswith('acq-'):
                            flag = 'trc-' + flag.split('-')[-1]
                        flags[i] = flag

                # fix subject and session names
                for i, flag in enumerate(flags):
                    if flag.startswith('sub-'):
                        flag = f'sub-{participant_id:04d}'
                    if flag.startswith(('ses-', 'sess-')):
                        flag = f'ses-{session_id:04d}'
                    flags[i] = flag

                basename = '_'.join(flags)

                write_from_buffer(
                    f.extractfile(niimember),
                    modality_path / f'{basename}.nii.gz',
                    logger=logger,
                )
                if jsonmember is None:
                    logger.warning(
                        f"No json found in {tarpath.name}"
                    )
                else:
                    write_from_buffer(
                        f.extractfile(jsonmember),
                        modality_path / f'{basename}.json',
                        logger=logger,
                    )
                if submodality_type == 'dwi':
                    if bvalmember is None:
                        logger.warning(f"No bval found in {tarpath.name}")
                    else:
                        write_from_buffer(
                            f.extractfile(bvalmember),
                            modality_path / f'{basename}.bval',
                            logger=logger,
                        )
                    if bvecmember is None:
                        logger.warning(f"No bvec found in {tarpath.name}")
                    else:
                        write_from_buffer(
                            f.extractfile(bvecmember),
                            modality_path / f'{basename}.bvec',
                            logger=logger,
                        )
                if submodality_type == 'pet':
                    if tsvmember is None:
                        logger.warning(f"No tsv found in {tarpath.name}")
                    else:
                        write_from_buffer(
                            f.extractfile(tsvmember),
                            modality_path / f'{basename}.tsv',
                            logger=logger,
                        )


def make_participants(path_xlsx, path_tsv):

    """
    oasis_header = [
        'SUB_ID',
        'MRI_ID',
        'Group',
        'Visit',
        'Delay',
        'M/F',
        'Hand',
        'Age',
        'Educ',
        'SES',
        'MMSE',
        'CDR',
        'eTIV',
        'nWBV',
        'ASF',
    ]
    """

    participants_header = [
        'participant_id',
        'sex',
        'handedness',
        'age',
    ]

    def iter_rows():
        xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
        xlsx = xlsx[xlsx.sheetnames[0]]
        yield participants_header
        for nrow in range(2, xlsx.max_row+1):
            id = xlsx[nrow][0].value
            visit = int(xlsx[nrow][3].value)
            if visit != 1:
                continue
            sex = xlsx[nrow][5].value
            hand = xlsx[nrow][6].value
            age = xlsx[nrow][7].value
            id = int(id.split('_')[-1])
            yield [f'sub-{id:04d}', sex, hand, age]

    write_tsv(iter_rows(), path_tsv)


def make_sessions(path_xlsx):

    """
    oasis_header = [
        'SUB_ID',
        'MRI_ID',
        'Group',
        'Visit',
        'Delay',
        'M/F',
        'Hand',
        'Age',
        'Educ',
        'SES',
        'MMSE',
        'CDR',
        'eTIV',
        'nWBV',
        'ASF',
    ]
    """

    sessions_header = [
        'session_id',
        'delay',
        'pathology',
        'age',
        'educ',
        'ses',
        'mmse',
        'cdr',
        'etiv',
        'nwbv',
        'asf',
    ]

    pathology_map = {
        "Nondemented": "N",
        "Demented": "D",
        "Converted": "C",
    }

    xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
    xlsx = xlsx[xlsx.sheetnames[0]]
    sessions_tables = {}
    for nrow in range(2, xlsx.max_row+1):
        id = int(xlsx[nrow][0].value.split('_')[-1])
        sessions_tables.setdefault(id, [sessions_header])
        sessions_tables[id].append([
            f'ses-{xlsx[nrow][3].value}',           # visit
            xlsx[nrow][4].value,                    # delay
            pathology_map[xlsx[nrow][2].value],     # pathology
            xlsx[nrow][7].value,                    # age
            xlsx[nrow][8].value,                    # educ
            xlsx[nrow][9].value,                    # ses
            xlsx[nrow][10].value,                   # mmse
            xlsx[nrow][11].value,                   # cdr
            xlsx[nrow][12].value,                   # tiv
            xlsx[nrow][13].value,                   # nwbv
            xlsx[nrow][14].value,                   # asf
        ])
    return sessions_tables


if __name__ == '__main__':

    root = logger.getLogger()
    root.setLevel(0)
    handler = logger.StreamHandler(sys.stdout)
    handler.setLevel(0)
    formatter = logger.Formatter('%(levelname)s | %(message)s')
    handler.setFormatter(formatter)
    root.addHandler(handler)

    bidsify()
