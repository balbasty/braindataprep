import os
import sys
import click
import tarfile
import logging
import nibabel
from braindataprep.utils import (
    get_tree_path,
    fileparts,
    nibabel_convert,
    write_tsv,
    write_from_buffer,
    copy_from_buffer,
    copy_json,
)

try:
    import openpyxl
except ImportError:
    logging.error(
        'Cannot find `openpyxl`. Did you install with [oasis2] flag? '
        'Try `pip install braindataprep[oasis2]`.'
    )

"""
Expected input
--------------
OASIS-II/
    sourcedata/
        OAS2_RAW_PART{1|2}.tar.gz
        oasis_longitudinal_demographics.xlsx

Expected output
---------------
OASIS-II/
    dataset_description.json
    participants.{tsv|json}
    sessions.json
    rawdata/
        sub-{04d}/
            sub-{04d}_sessions.tsv
            ses-{d}/
                anat/
                    sub-{04d}_run-{d}_T1w.{nii.gz|json}
"""

"""Folder containing template README/JSON/..."""
TPLDIR = os.path.join(os.path.dirname(__file__), 'templates')


@click.command()
@click.option(
    '--path', default=None, help='Path to tree')
@click.option(
    '--key', multiple=True,
    type=click.Choice(["meta", "raw"]),
    help='Only download these keys')
@click.option(
    '--part', multiple=True, type=int,
    help='Only download these parts (1, 2)')
@click.option(
    '--sub', multiple=True, type=int,
    help='Only download these subjects')
@click.option(
    '--json-only', is_flag=True, default=False,
    help='Only write jsons (not volumes)'
)
def bidsify(path, key, part, sub, json_only):
    logging.info('OASIS-I - bidsify')
    path = get_tree_path(path)
    keys = set(key or ['meta', 'raw'])
    parts = set(part or [1, 2])
    subs = sub
    oasispath = os.path.join(path, 'OASIS-2')
    src = os.path.join(oasispath, 'sourcedata')
    raw = os.path.join(oasispath, 'rawdata')
    if not os.path.exists(src):
        raise FileNotFoundError('sourcedata folder not found')

    # ------------------------------------------------------------------
    #   Write toplevel meta-information
    #    - README
    #    - dataset_description.json
    #    - participants.tsv
    #    - participants.json
    # ------------------------------------------------------------------
    if 'meta' in keys:

        copy_from_buffer(
            os.path.join(TPLDIR, 'README'),
            os.path.join(oasispath, 'README')
        )

        copy_json(
            os.path.join(TPLDIR, 'dataset_description.json'),
            os.path.join(oasispath, 'dataset_description.json')
        )

        copy_json(
            os.path.join(TPLDIR, 'participants.json'),
            os.path.join(oasispath, 'participants.json')
        )

        make_participants(
            os.path.join(src, 'oasis_longitudinal_demographics.xlsx'),
            os.path.join(oasispath, 'participants.tsv'),
        )

        copy_json(
            os.path.join(TPLDIR, 'sessions.json'),
            os.path.join(oasispath, 'sessions.json')
        )

    # we always need this to write subject-specific session files
    session_tables = make_sessions(
        os.path.join(src, 'oasis_longitudinal_demographics.xlsx')
    )

    if not keys.intersection(set(['raw'])):
        return

    # ------------------------------------------------------------------
    #   Write toplevel meta-information
    #    - README
    #    - dataset_description.json
    #    - participants.tsv
    #    - participants.json
    #    - sessions.json
    # ------------------------------------------------------------------
    for part in parts:

        # --------------------------------------------------------------
        #   Convert raw and minimally processed data
        #   from "oasis_cross-sectional_disc{d}.tar.gz"
        # --------------------------------------------------------------

        logging.info(f'process part {part}')
        tarpath = os.path.join(src, f'OAS2_RAW_PART{part}.tar.gz')
        if not os.path.exists(tarpath):
            logging.warning(f'OAS2_RAW_PART{part}.tar.gz not found')
            continue

        with tarfile.open(tarpath, 'r:gz') as f:
            # ----------------------------------------------------------
            #   Save all subject ids and runs contained in this part
            # ----------------------------------------------------------
            subjects = {}
            for member in f.getmembers():
                if 'RAW' not in member.name:
                    continue
                if not member.name.endswith('.img'):
                    continue
                dirname, basename, _ = fileparts(member.name)
                id = os.path.basename(os.path.dirname(dirname))
                _, id, ses = id.split('_')
                id, ses = int(id), int(ses[2:])
                if not basename.startswith('mpr'):
                    # Found a folder with a weird file 3906-3.nift.img
                    # which seems to be a duplicate of mpr-1.nifti.img
                    # Let's skip it
                    continue
                run = int(basename[4])
                if subs and id not in subs:
                    continue
                subjects.setdefault(id, {})
                subjects[id].setdefault(ses, [])
                subjects[id][ses].append(run)

            # ----------------------------------------------------------
            #   Convert raw data
            #   (per-run scans only, the average is a derivative)
            # ----------------------------------------------------------
            for id, sessions in subjects.items():
                subdir = os.path.join(raw, f'sub-{id:04d}')
                write_tsv(
                    session_tables[id],
                    os.path.join(subdir, f'sub-{id:04d}_sessions.tsv')
                )
                for ses, runs in sessions.items():
                    sesdir = os.path.join(subdir, f'ses-{ses}', 'anat')
                    oasisraw = (
                        f'OAS2_RAW_PART{part}/'
                        f'OAS2_{id:04d}_MR{ses}/'
                        f'RAW'
                    )
                    for run in runs:
                        hdrname = f'sub-{id:04d}_run-{run:d}_T1w.hdr'
                        imgname = f'sub-{id:04d}_run-{run:d}_T1w.img'
                        scanname = f'sub-{id:04d}_run-{run:d}_T1w.nii.gz'
                        jsonname = f'sub-{id:04d}_run-{run:d}_T1w.json'
                        copy_json(
                            os.path.join(TPLDIR, 'T1w.json'),
                            os.path.join(sesdir, jsonname)
                        )
                        if json_only:
                            continue
                        write_from_buffer(
                            f.extractfile(f'{oasisraw}/mpr-{run}.nifti.hdr'),
                            os.path.join(sesdir, hdrname)
                        )
                        write_from_buffer(
                            f.extractfile(f'{oasisraw}/mpr-{run}.nifti.img'),
                            os.path.join(sesdir, imgname)
                        )
                        nibabel_convert(
                            os.path.join(sesdir, imgname),
                            os.path.join(sesdir, scanname),
                            inp_format=nibabel.AnalyzeImage,
                            remove=True,
                        )


def make_participants(path_xlsx, path_tsv):

    """
    oasis_header = [
        'SUB_ID',
        'MRI_ID',
        'Group',
        'Visit',
        'Delay',
        'M/F',
        'Hand',
        'Age',
        'Educ',
        'SES',
        'MMSE',
        'CDR',
        'eTIV',
        'nWBV',
        'ASF',
    ]
    """

    participants_header = [
        'participant_id',
        'sex',
        'handedness',
        'age',
    ]

    def iter_rows():
        xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
        xlsx = xlsx[xlsx.sheetnames[0]]
        yield participants_header
        for nrow in range(2, xlsx.max_row+1):
            id = xlsx[nrow][0].value
            visit = int(xlsx[nrow][3].value)
            if visit != 1:
                continue
            sex = xlsx[nrow][5].value
            hand = xlsx[nrow][6].value
            age = xlsx[nrow][7].value
            id = int(id.split('_')[-1])
            yield [f'sub-{id:04d}', sex, hand, age]

    write_tsv(iter_rows(), path_tsv)


def make_sessions(path_xlsx):

    """
    oasis_header = [
        'SUB_ID',
        'MRI_ID',
        'Group',
        'Visit',
        'Delay',
        'M/F',
        'Hand',
        'Age',
        'Educ',
        'SES',
        'MMSE',
        'CDR',
        'eTIV',
        'nWBV',
        'ASF',
    ]
    """

    sessions_header = [
        'session_id',
        'delay',
        'pathology',
        'age',
        'educ',
        'ses',
        'mmse',
        'cdr',
        'etiv',
        'nwbv',
        'asf',
    ]

    pathology_map = {
        "Nondemented": "N",
        "Demented": "D",
        "Converted": "C",
    }

    xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
    xlsx = xlsx[xlsx.sheetnames[0]]
    sessions_tables = {}
    for nrow in range(2, xlsx.max_row+1):
        id = int(xlsx[nrow][0].value.split('_')[-1])
        sessions_tables.setdefault(id, [sessions_header])
        sessions_tables[id].append([
            f'ses-{xlsx[nrow][3].value}',           # visit
            xlsx[nrow][4].value,                    # delay
            pathology_map[xlsx[nrow][2].value],     # pathology
            xlsx[nrow][7].value,                    # age
            xlsx[nrow][8].value,                    # educ
            xlsx[nrow][9].value,                    # ses
            xlsx[nrow][10].value,                   # mmse
            xlsx[nrow][11].value,                   # cdr
            xlsx[nrow][12].value,                   # tiv
            xlsx[nrow][13].value,                   # nwbv
            xlsx[nrow][14].value,                   # asf
        ])
    return sessions_tables


if __name__ == '__main__':

    root = logging.getLogger()
    root.setLevel(0)
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(0)
    formatter = logging.Formatter('%(levelname)s | %(message)s')
    handler.setFormatter(formatter)
    root.addHandler(handler)

    bidsify()
